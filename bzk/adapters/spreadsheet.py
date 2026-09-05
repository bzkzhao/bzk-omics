"""Workbook reading — the one home for turning an `.xlsx` sheet into rows.

**Why a module and not a block in each caller.** `bzk/sources/protein_groups.py` carried the only
`openpyxl` block in the tree, and `bzk/adapters/perseus.py` now needs one too: the analysis-output
export it was written for also arrives as a spreadsheet. Two blocks reading one format is the shape
`bzk/adapters/maxquant.py` exists to refuse — the spill-line guard was moved there rather than
copied, on the ground that *any* MaxQuant reader hits it. The same argument holds one format over,
so the reader lives beside that one rather than inside either caller.

**Raw values out, and text as a second function rather than a flag.** `rows` yields cells exactly as
`openpyxl` does, which is what `bzk/sources/protein_groups.py` was already reading — that module's
own splitter depends on `str(None)` arriving as `'None'`, and a reader that helpfully blanked it
would change a measurement while looking like a refactor. `text_rows` is what the adapter wants: a
string per cell, the empty string for a missing one. Two functions, so neither caller can silently
receive the other's reading.

**`load_workbook` is called with `read_only=True` and nothing else**, which is exactly what
`bzk/sources/protein_groups.py` called before this module existed. `data_only` is deliberately not
passed: it changes which value a formula cell yields, and a reader extracted to remove a duplicate
is not the place to change what a caller reads.

**`openpyxl` is imported inside the function, and that is not a style choice.** It is pinned in
`pyproject.toml`'s **dev** dependency group rather than the project's, and importing it at module
scope would make `bzk/adapters/perseus.py` — a core ingestion path — unimportable in an install
without that group. Inside the function, a missing reader is a clear error raised at the point of
reading rather than an `ImportError` at load time.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

#: The first four bytes of a ZIP container, which is what an `.xlsx` is. A caller has to know which
#: reader to use before it knows anything else about the file, and the extension is not evidence —
#: `ARCHITECTURE.md` §3's *content, not name*, which is the same ground `sniff` already stands on.
ZIP_MAGIC = b"PK\x03\x04"


class SpreadsheetError(ValueError):
    """A workbook cannot be read. Never downgraded to a warning (`CLAUDE.md`)."""


def looks_like_a_workbook(raw: bytes) -> bool:
    """True for bytes that open as a ZIP container.

    A container test, not a validity test: it says which reader to try, and `rows` says whether the
    bytes are really a workbook. Splitting it that way keeps the caller's dispatch total — every
    byte string goes to exactly one reader — while still refusing loudly on a truncated file.
    """
    return raw.startswith(ZIP_MAGIC)


def _load(source: Path | bytes, *, read_only: bool) -> Any:
    """The workbook, with the `openpyxl` import and the failure wrapping in one place."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - openpyxl is pinned in the dev group
        raise SpreadsheetError(
            "openpyxl is not installed, so no workbook can be read; it is pinned in "
            "pyproject.toml's dev dependency group"
        ) from exc
    handle: Any = io.BytesIO(source) if isinstance(source, bytes) else source
    try:
        return openpyxl.load_workbook(handle, read_only=read_only)
    except Exception as exc:
        raise SpreadsheetError(f"not a readable workbook: {type(exc).__name__}: {exc}") from exc


def rows(source: Path | bytes, *, min_row: int = 1) -> list[tuple[Any, ...]]:
    """Every cell of the workbook's **first** sheet, as `openpyxl` yields it.

    First sheet and not a named one: both callers read single-sheet exports, and a name would be a
    convention neither file states. `min_row` is 1-based and matches `openpyxl`'s own, so a caller
    skipping a title row asks for what it means rather than slicing afterwards.
    """
    workbook = _load(source, read_only=True)
    try:
        sheet = workbook.worksheets[0]
        return list(sheet.iter_rows(min_row=min_row, values_only=True))
    finally:
        workbook.close()


def merged_spans(source: Path | bytes) -> list[tuple[int, int, int, int]]:
    """The first sheet's merged ranges, as 0-based inclusive `(top, left, bottom, right)`.

    **A separate read, because the streaming one cannot answer.** A worksheet opened with
    `read_only=True` has no `merged_cells` at all — measured: touching it raises
    `AttributeError: 'ReadOnlyWorksheet' object has no attribute 'merged_cells'` — so the ranges
    cannot come from the open that `rows` does. Opened without the flag they are there, and the cell
    *values* are identical under both flags, which is what makes reading the ranges separately safe
    for a caller that reads only values.

    **A function rather than a flag on `rows`, for the reason this module is already split in two.**
    `bzk/sources/protein_groups.py` calls `rows` and its measurement is pinned; nothing it calls
    changes here, and it never calls this.

    **The cost, stated rather than argued away.** The workbook is opened a second time and that open
    is not streaming, so the first sheet is materialised in full. The alternative — reading
    `xl/worksheets/*.xml` out of the container by hand — would reimplement a pinned library's
    container parse to save a read, which buys a new failure surface for a saving nobody has
    measured a need for.

    Sorted, because `openpyxl` holds the ranges in a set and their order is otherwise arbitrary.
    """
    workbook = _load(source, read_only=False)
    try:
        sheet = workbook.worksheets[0]
        return sorted(
            (r.min_row - 1, r.min_col - 1, r.max_row - 1, r.max_col - 1)
            for r in sheet.merged_cells.ranges
        )
    finally:
        workbook.close()


def text_rows(source: Path | bytes, *, min_row: int = 1) -> list[list[str]]:
    """`rows`, with every cell as text and a missing cell as the empty string.

    The empty string rather than `'None'`: a caller reading header cells needs *absent* to be
    distinguishable from a cell whose text is the word, and `str(None)` erases that difference.
    """
    return [["" if c is None else str(c) for c in row] for row in rows(source, min_row=min_row)]
